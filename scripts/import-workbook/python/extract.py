#!/usr/bin/env python3
"""Extract one JSONL record per non-empty day cell of the source workbook.

    uv run --project scripts/import-workbook python scripts/import-workbook/python/extract.py

Writes ``data/staging/cells.jsonl`` -- exactly one record per non-empty day
cell, ordered by ``(row, col)``, with these keys and no others::

    sheet, row, col, day_index, week_label, week_number, local_date,
    raw_text, raw_text_sha256, workbook_sha256

``raw_text`` is the cell value *verbatim*. Nothing is normalised here; that
happens later, in TypeScript, where the transform log stays reconstructible.

This module also owns how the workbook is read, and ``inspect_workbook.py``
imports it. That is deliberate: the profile report claims 170 non-empty cells
and this script writes 170 records, and those two numbers are only a meaningful
gate if both come from the same reader.

Exits non-zero -- writing nothing -- if the cell count is not the expected 170
or if any of the 53 week labels disagrees with the anchor computation.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import workbook_dates as wd

EXPECTED_CELL_COUNT = 170
DEFAULT_OUTPUT_RELPATH = "data/staging/cells.jsonl"
SOURCE_DIR_RELPATH = "data/source"


def repo_root() -> Path:
    """The repo root, resolved from this file -- never from the cwd."""
    return Path(__file__).resolve().parents[3]


def default_workbook_path() -> Path:
    """The single ``.xlsx`` under ``data/source/`` (pinned by MANIFEST.sha256)."""
    source_dir = repo_root() / SOURCE_DIR_RELPATH
    candidates = sorted(p for p in source_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not candidates:
        raise SystemExit(
            f"No .xlsx found in {source_dir}. The workbook is personal health data and is "
            f"gitignored -- copy it there before running the importer."
        )
    if len(candidates) > 1:
        names = ", ".join(p.name for p in candidates)
        raise SystemExit(
            f"Expected exactly one .xlsx in {source_dir}, found {len(candidates)}: {names}"
        )
    return candidates[0]


def _rel(path: Path, root: Path) -> str:
    """Repo-relative when possible, absolute otherwise -- only for display."""
    try:
        return str(path.resolve().relative_to(root))
    except ValueError:
        return str(path)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


@dataclass(frozen=True)
class DayCell:
    """One non-empty day cell, still completely unparsed."""

    row: int
    col: int
    raw_text: str


@dataclass(frozen=True)
class Workbook:
    path: Path
    sha256: str
    sheet_name: str
    dimensions: str
    max_row: int
    max_column: int
    header: tuple[Any, ...]
    #: row number -> column A text, for every row that has one.
    week_labels: dict[int, str]
    #: Every non-empty day cell (columns B..H), ordered by (row, col).
    day_cells: tuple[DayCell, ...]
    #: Rows in the day-column range that hold no data at all.
    empty_rows: tuple[int, ...]


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def read_workbook(path: Path) -> Workbook:
    """Read the workbook into plain data. The only place openpyxl is touched."""
    wb = load_workbook(path, data_only=True)
    if wd.SHEET_NAME not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {wd.SHEET_NAME!r} not found in {path.name}; sheets are {wb.sheetnames!r}"
        )
    ws = wb[wd.SHEET_NAME]

    rows = list(
        ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
            values_only=True,
        )
    )
    header = tuple(rows[0]) if rows else ()

    week_labels: dict[int, str] = {}
    day_cells: list[DayCell] = []
    empty_rows: list[int] = []

    for row_number, row in enumerate(rows[1:], start=2):
        label = row[0] if row else None
        if not _is_empty(label):
            week_labels[row_number] = str(label)

        row_has_data = False
        for col in range(wd.FIRST_DAY_COL, wd.LAST_DAY_COL + 1):
            value = row[col - 1] if col - 1 < len(row) else None
            if _is_empty(value):
                continue
            row_has_data = True
            text = value if isinstance(value, str) else str(value)
            day_cells.append(DayCell(row=row_number, col=col, raw_text=text))
        if not row_has_data:
            empty_rows.append(row_number)

    day_cells.sort(key=lambda c: (c.row, c.col))
    wb.close()

    return Workbook(
        path=path,
        sha256=sha256_file(path),
        sheet_name=ws.title,
        dimensions=ws.calculate_dimension(),
        max_row=ws.max_row,
        max_column=ws.max_column,
        header=header,
        week_labels=week_labels,
        day_cells=tuple(day_cells),
        empty_rows=tuple(empty_rows),
    )


def check_all_week_labels(week_labels: dict[int, str]) -> list[str]:
    """Cross-check every label against the anchor. Returns a list of problems."""
    problems: list[str] = []
    for row_number in sorted(week_labels):
        label = week_labels[row_number]
        check = wd.check_week_label(label)
        expected_week = wd.week_number_for_row(row_number)
        if check.week_number is not None and check.week_number != expected_week:
            problems.append(
                f"R{row_number}: label says week {check.week_number}, row implies week {expected_week}"
            )
        problems.extend(f"R{row_number}: {p}" for p in check.problems)
    return problems


def build_records(book: Workbook) -> list[dict[str, Any]]:
    """One record per non-empty day cell. Key order is part of the contract."""
    records: list[dict[str, Any]] = []
    for cell in book.day_cells:
        week_number = wd.week_number_for_row(cell.row)
        records.append(
            {
                "sheet": book.sheet_name,
                "row": cell.row,
                "col": cell.col,
                "day_index": wd.day_index_for_col(cell.col),
                "week_label": book.week_labels.get(cell.row, ""),
                "week_number": week_number,
                "local_date": wd.cell_local_date(cell.row, cell.col),
                "raw_text": cell.raw_text,
                "raw_text_sha256": sha256_text(cell.raw_text),
                "workbook_sha256": book.sha256,
            }
        )
    return records


def write_jsonl(records: list[dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="\n") as fh:
        for record in records:
            fh.write(json.dumps(record, ensure_ascii=False) + "\n")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--workbook", type=Path, default=None, help="Override the source .xlsx path")
    parser.add_argument("--out", type=Path, default=None, help=f"Override {DEFAULT_OUTPUT_RELPATH}")
    parser.add_argument(
        "--expect-count",
        type=int,
        default=EXPECTED_CELL_COUNT,
        help=f"Fail unless this many cells are found (default {EXPECTED_CELL_COUNT})",
    )
    args = parser.parse_args(argv)

    root = repo_root()
    workbook_path = args.workbook or default_workbook_path()
    out_path = args.out or (root / DEFAULT_OUTPUT_RELPATH)

    book = read_workbook(workbook_path)

    failures: list[str] = []

    label_problems = check_all_week_labels(book.week_labels)
    if label_problems:
        failures.append(
            "Week labels disagree with the anchor computation "
            f"({wd.WEEK_ANCHOR_ISO} + 7 x (week - 1)):\n  " + "\n  ".join(label_problems)
        )

    out_of_range = sorted(
        {c.row for c in book.day_cells if not wd.FIRST_DATA_ROW <= c.row <= wd.LAST_LABEL_ROW}
    )
    if out_of_range:
        failures.append(
            f"Day cells found outside rows {wd.FIRST_DATA_ROW}..{wd.LAST_LABEL_ROW}: {out_of_range}"
        )

    missing_labels = sorted({c.row for c in book.day_cells if c.row not in book.week_labels})
    if missing_labels:
        failures.append(f"Rows with data but no week label in column A: {missing_labels}")

    if len(book.day_cells) != args.expect_count:
        failures.append(
            f"Expected {args.expect_count} non-empty day cells, found {len(book.day_cells)}. "
            "The workbook changed shape -- re-run inspect_workbook.py and review the profile "
            "before importing."
        )

    if failures:
        print("extract: FAILED, nothing written.", file=sys.stderr)
        for failure in failures:
            print(f"  - {failure}", file=sys.stderr)
        return 1

    records = build_records(book)
    write_jsonl(records, out_path)

    rows_with_data = sorted({r["row"] for r in records})
    dates = [r["local_date"] for r in records]

    label_count = len(book.week_labels)
    beyond = sum(1 for r in records if r["row"] > wd.LAST_DATA_ROW)

    print(f"workbook       {_rel(workbook_path, root)}")
    print(f"sha256         {book.sha256}")
    print(f"sheet          {book.sheet_name} ({book.dimensions})")
    print(f"week labels    {label_count}/{label_count} match the anchor computation")
    print(f"records        {len(records)}")
    print(f"rows with data {rows_with_data[0]}..{rows_with_data[-1]} ({len(rows_with_data)} weeks)")
    print(f"rows {wd.LAST_DATA_ROW + 1}..{wd.LAST_LABEL_ROW}    {beyond} records")
    print(f"date range     {min(dates)} .. {max(dates)}")
    print(f"wrote          {_rel(out_path, root)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
